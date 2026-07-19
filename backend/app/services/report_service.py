import csv
import io
import logging
import uuid
from datetime import datetime, timedelta, timezone

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.models.application_tracking import Application, ApplicationTagMapping
from app.schemas.dashboard import StatusCount, TopCompany

logger = logging.getLogger(__name__)


class ReportService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def generate_report(
        self,
        user_id: uuid.UUID,
        report_type: str,
        fmt: str,
        date: str | None = None,
    ) -> tuple[bytes, str, str]:
        now = datetime.now(timezone.utc)
        if report_type == "daily":
            if date:
                period_start = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            else:
                period_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            period_end = period_start + timedelta(days=1)
            filename_prefix = f"daily_report_{period_start.strftime('%Y-%m-%d')}"
        elif report_type == "weekly":
            if date:
                period_start = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            else:
                period_start = (now - timedelta(days=now.weekday())).replace(
                    hour=0, minute=0, second=0, microsecond=0
                )
            period_end = period_start + timedelta(days=7)
            filename_prefix = f"weekly_report_{period_start.strftime('%Y-%m-%d')}"
        elif report_type == "monthly":
            if date:
                period_start = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            else:
                period_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if period_start.month == 12:
                period_end = period_start.replace(year=period_start.year + 1, month=1)
            else:
                period_end = period_start.replace(month=period_start.month + 1)
            filename_prefix = f"monthly_report_{period_start.strftime('%Y-%m')}"
        else:
            raise ValueError(f"Unsupported report type: {report_type}")

        apps = await self._get_applications_in_range(user_id, period_start, period_end)
        total = len(apps)
        status_breakdown = await self._get_status_breakdown(user_id, period_start, period_end)
        top_companies = await self._get_top_companies(user_id, period_start, period_end, limit=10)
        interview_rate, success_rate = await self._get_rates(user_id, period_start, period_end)
        daily_breakdown = await self._get_daily_breakdown(user_id, period_start, period_end)

        if fmt == "csv":
            content, filename = self._build_csv(
                apps, filename_prefix, report_type, period_start, period_end,
                total, status_breakdown, top_companies, interview_rate, success_rate,
                daily_breakdown,
            )
            return content, filename, "text/csv"
        elif fmt == "xlsx":
            content, filename = self._build_xlsx(
                apps, filename_prefix, report_type, period_start, period_end,
                total, status_breakdown, top_companies, interview_rate, success_rate,
                daily_breakdown,
            )
            return content, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif fmt == "pdf":
            content, filename = self._build_pdf(
                apps, filename_prefix, report_type, period_start, period_end,
                total, status_breakdown, top_companies, interview_rate, success_rate,
                daily_breakdown,
            )
            return content, filename, "application/pdf"
        else:
            raise ValueError(f"Unsupported format: {fmt}")

    async def _get_applications_in_range(
        self, user_id: uuid.UUID, start: datetime, end: datetime,
    ) -> list[Application]:
        stmt = (
            select(Application)
            .options(joinedload(Application.tag_mappings).joinedload(ApplicationTagMapping.tag))
            .where(
                Application.user_id == user_id,
                Application.created_at >= start,
                Application.created_at < end,
            )
            .order_by(Application.created_at.desc())
        )
        result = await self.session.execute(stmt)
        return list(result.unique().scalars().all())

    async def _get_status_breakdown(
        self, user_id: uuid.UUID, start: datetime, end: datetime,
    ) -> list[StatusCount]:
        result = await self.session.execute(
            select(Application.status, func.count(Application.id).label("cnt"))
            .where(
                Application.user_id == user_id,
                Application.created_at >= start,
                Application.created_at < end,
            )
            .group_by(Application.status)
        )
        return [StatusCount(status=row[0], count=row[1]) for row in result]

    async def _get_top_companies(
        self, user_id: uuid.UUID, start: datetime, end: datetime, limit: int = 10,
    ) -> list[TopCompany]:
        result = await self.session.execute(
            select(Application.company_name, func.count(Application.id).label("cnt"))
            .where(
                Application.user_id == user_id,
                Application.created_at >= start,
                Application.created_at < end,
            )
            .group_by(Application.company_name)
            .order_by(func.count(Application.id).desc())
            .limit(limit)
        )
        return [TopCompany(company_name=row[0], count=row[1]) for row in result]

    async def _get_rates(
        self, user_id: uuid.UUID, start: datetime, end: datetime,
    ) -> tuple[float, float]:
        total_result = await self.session.execute(
            select(func.count(Application.id)).where(
                Application.user_id == user_id,
                Application.created_at >= start,
                Application.created_at < end,
            )
        )
        total = total_result.scalar() or 0
        interview_result = await self.session.execute(
            select(func.count(Application.id)).where(
                Application.user_id == user_id,
                Application.status.in_(["interview", "offer", "accepted"]),
                Application.created_at >= start,
                Application.created_at < end,
            )
        )
        interview_count = interview_result.scalar() or 0
        accepted_result = await self.session.execute(
            select(func.count(Application.id)).where(
                Application.user_id == user_id,
                Application.status == "accepted",
                Application.created_at >= start,
                Application.created_at < end,
            )
        )
        accepted_count = accepted_result.scalar() or 0
        interview_rate = round((interview_count / total * 100), 1) if total > 0 else 0.0
        success_rate = round((accepted_count / total * 100), 1) if total > 0 else 0.0
        return interview_rate, success_rate

    async def _get_daily_breakdown(
        self, user_id: uuid.UUID, start: datetime, end: datetime,
    ) -> list[dict]:
        rows = await self.session.execute(
            select(func.date(Application.created_at).label("d"), func.count(Application.id).label("cnt"))
            .where(
                Application.user_id == user_id,
                Application.created_at >= start,
                Application.created_at < end,
            )
            .group_by(func.date(Application.created_at))
            .order_by(func.date(Application.created_at))
        )
        return [{"date": str(row[0]), "count": row[1]} for row in rows]

    def _build_csv(
        self, apps, filename_prefix, report_type,
        period_start, period_end, total, status_breakdown,
        top_companies, interview_rate, success_rate, daily_breakdown,
    ) -> tuple[bytes, str]:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"{report_type.title()} Report"])
        writer.writerow([f"Period: {period_start.date()} to {period_end.date()}"])
        writer.writerow([])
        writer.writerow(["Summary"])
        writer.writerow(["Total Applications", total])
        writer.writerow(["Interview Rate", f"{interview_rate}%"])
        writer.writerow(["Success Rate", f"{success_rate}%"])
        writer.writerow([])
        writer.writerow(["Status Breakdown"])
        writer.writerow(["Status", "Count"])
        for sb in status_breakdown:
            writer.writerow([sb.status, sb.count])
        writer.writerow([])
        if daily_breakdown:
            writer.writerow(["Daily Breakdown"])
            writer.writerow(["Date", "Applications"])
            for d in daily_breakdown:
                writer.writerow([d["date"], d["count"]])
            writer.writerow([])
        writer.writerow(["Top Companies"])
        writer.writerow(["Company", "Applications"])
        for tc in top_companies:
            writer.writerow([tc.company_name, tc.count])
        writer.writerow([])
        writer.writerow(["Application Details"])
        writer.writerow(["ID", "Job Title", "Company", "Status", "Location", "Applied At", "Tags"])
        for app in apps:
            tags_str = ", ".join(m.tag.name for m in app.tag_mappings) if app.tag_mappings else ""
            writer.writerow([
                str(app.id), app.job_title, app.company_name, app.status,
                app.location or "", app.applied_at.isoformat() if app.applied_at else "",
                tags_str,
            ])
        content = output.getvalue().encode("utf-8")
        filename = f"{filename_prefix}.csv"
        return content, filename

    def _build_xlsx(
        self, apps, filename_prefix, report_type,
        period_start, period_end, total, status_breakdown,
        top_companies, interview_rate, success_rate, daily_breakdown,
    ) -> tuple[bytes, str]:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append([f"{report_type.title()} Report"])
        ws.append([f"Period: {period_start.date()} to {period_end.date()}"])
        ws.append([])
        ws.append(["Summary"])
        ws.append(["Total Applications", total])
        ws.append(["Interview Rate", f"{interview_rate}%"])
        ws.append(["Success Rate", f"{success_rate}%"])
        ws.append([])
        ws.append(["Status Breakdown"])
        ws.append(["Status", "Count"])
        for sb in status_breakdown:
            ws.append([sb.status, sb.count])
        if daily_breakdown:
            ws.append([])
            ws2 = wb.create_sheet("Daily Breakdown")
            ws2.append(["Date", "Applications"])
            for d in daily_breakdown:
                ws2.append([d["date"], d["count"]])
        ws.append([])
        ws3 = wb.create_sheet("Top Companies")
        ws3.append(["Company", "Applications"])
        for tc in top_companies:
            ws3.append([tc.company_name, tc.count])
        ws4 = wb.create_sheet("Application Details")
        ws4.append(["ID", "Job Title", "Company", "Status", "Location", "Applied At", "Tags"])
        for app in apps:
            tags_str = ", ".join(m.tag.name for m in app.tag_mappings) if app.tag_mappings else ""
            ws4.append([
                str(app.id), app.job_title, app.company_name, app.status,
                app.location or "", app.applied_at.isoformat() if app.applied_at else "",
                tags_str,
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"{filename_prefix}.xlsx"
        return output.getvalue(), filename

    def _build_pdf(
        self, apps, filename_prefix, report_type,
        period_start, period_end, total, status_breakdown,
        top_companies, interview_rate, success_rate, daily_breakdown,
    ) -> tuple[bytes, str]:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"{report_type.title()} Report", styles["Title"]))
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph(f"Period: {period_start.date()} to {period_end.date()}", styles["Normal"]))
        elements.append(Spacer(1, 0.2 * inch))

        elements.append(Paragraph("Summary", styles["Heading2"]))
        summary_data = [
            ["Total Applications", str(total)],
            ["Interview Rate", f"{interview_rate}%"],
            ["Success Rate", f"{success_rate}%"],
        ]
        t = Table(summary_data, colWidths=[2 * inch, 1 * inch])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 0.2 * inch))

        elements.append(Paragraph("Status Breakdown", styles["Heading2"]))
        status_data = [["Status", "Count"]]
        status_data.extend([[sb.status, str(sb.count)] for sb in status_breakdown])
        t2 = Table(status_data, colWidths=[2 * inch, 1 * inch])
        t2.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ]))
        elements.append(t2)
        elements.append(Spacer(1, 0.2 * inch))

        if daily_breakdown:
            elements.append(Paragraph("Daily Breakdown", styles["Heading2"]))
            daily_data = [["Date", "Applications"]]
            daily_data.extend([[d["date"], str(d["count"])] for d in daily_breakdown])
            t3 = Table(daily_data, colWidths=[2 * inch, 1 * inch])
            t3.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ]))
            elements.append(t3)
            elements.append(Spacer(1, 0.2 * inch))

        elements.append(Paragraph("Top Companies", styles["Heading2"]))
        company_data = [["Company", "Applications"]]
        company_data.extend([[tc.company_name, str(tc.count)] for tc in top_companies])
        t4 = Table(company_data, colWidths=[3 * inch, 1 * inch])
        t4.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ]))
        elements.append(t4)
        elements.append(Spacer(1, 0.3 * inch))

        elements.append(Paragraph(f"Total applications in this period: {total}", styles["Normal"]))

        doc.build(elements)
        buf.seek(0)
        filename = f"{filename_prefix}.pdf"
        return buf.getvalue(), filename
